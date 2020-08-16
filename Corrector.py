from openpyxl import Workbook,load_workbook
from openpyxl.styles import Alignment
from itertools import islice


wb2 = load_workbook("file.xlsx")
working_sheet = wb2.active
for cells in islice(working_sheet.iter_rows(min_col=3,max_col=4),1,None,2):
    for cell in cells:

        value = cell.value
        if value is not None:
            row = cell.row
            if cell.column == 3:
                to_cell = working_sheet["D"+str(row-1)]
                to_cell.value = "$"+str(value)
                to_cell.alignment = Alignment(horizontal='right')
                working_sheet["C"+str(row)] = None
            elif cell.column == 4:
                to_cell = working_sheet["E"+str(row-1)]
                to_cell.value = "$" + str(value)
                to_cell.alignment = Alignment(horizontal='right')
                working_sheet["D" + str(row)] = None
            else:
                raise Exception

for cells in islice(working_sheet.iter_rows(min_col=2,max_col=2),1,None,2):
    cell = cells[0]
    value = cell.value
    if value is not None:
        row = cell.row
        to_cell = working_sheet["C"+str(row-1)]
        to_cell.value = "$" + str(value)
        to_cell.alignment = Alignment(horizontal='right')
        working_sheet["B"+str(row)] = None

wb2.save("to_file.xlsx")
